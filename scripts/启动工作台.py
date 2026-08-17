# -*- coding: utf-8 -*-
"""
团支书工作台 - 桌面应用启动器
基于 pywebview 的 Windows 桌面软件
"""
import os
import sys
import json
import webbrowser
import threading
import urllib.request
from datetime import datetime

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 数据文件路径（与程序同目录）
APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_FILE = os.path.join(APP_DIR, "tuanzhishu_data.json")
THEME_FILE = os.path.join(APP_DIR, ".theme")


def load_theme():
    """读取主题标记：night=墨夜深色，其余=宣纸浅色（默认）"""
    try:
        if os.path.exists(THEME_FILE):
            with open(THEME_FILE, "r", encoding="utf-8") as f:
                if f.read().strip().lower() == "night":
                    return "night"
    except Exception:
        pass
    return "paper"


def window_bg_color():
    """pywebview 窗口底色跟随主题：深色用 #2A2620 避免启动瞬间闪浅色，浅色保持 #faf7f8"""
    return "#2A2620" if load_theme() == "night" else "#faf7f8"


def _norm_join_date(val):
    """入团时间规范化 → 'YYYY-MM'。兼容：日期对象 / Excel序列号 / 2024年12月 / 2024.12 / 2024/12 / 202412 / 2024-12。
    脏格式不丢数据：团龄/推优核查都依赖这个正则，格式乱则核查全歪（帝王标的雷位）。"""
    import re as _re
    import datetime as _dt
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m")
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        n = int(val)
        # YYYYMM 语义优先（如 202412 → 2024-12）
        if 190001 <= n <= 210012:
            return "%d-%02d" % (n // 100, n % 100)
        # Excel 日期序列号（1899-12-30 起算，与 openpyxl 一致）
        if 1 <= n <= 80000:
            try:
                return (_dt.date(1899, 12, 30) + _dt.timedelta(days=n)).strftime("%Y-%m")
            except Exception:
                return ""
        return ""
    s = str(val).strip()
    if not s:
        return ""
    m = _re.search(r"(\d{4})\D{0,2}(\d{1,2})", s)
    if m:
        return "%s-%02d" % (m.group(1), int(m.group(2)))
    return s[:7] if len(s) >= 7 else s


class Api:
    """暴露给前端 JS 的接口"""

    def save_data(self, json_str):
        """保存数据到本地 JSON 文件"""
        try:
            with open(DATA_FILE, "w", encoding="utf-8") as f:
                f.write(json_str)
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def load_data(self):
        """从本地 JSON 文件读取数据"""
        try:
            if os.path.exists(DATA_FILE):
                with open(DATA_FILE, "r", encoding="utf-8") as f:
                    return f.read()
            return None
        except Exception as e:
            return None

    def save_theme(self, theme):
        """前端切换主题时同步写标记，下次启动据此决定窗口底色（根治启动闪白）"""
        try:
            with open(THEME_FILE, "w", encoding="utf-8") as f:
                f.write("night" if theme == "night" else "paper")
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def get_backup_dir(self):
        """获取备份目录信息"""
        return APP_DIR

    def choose_backup_dir(self):
        """弹原生目录选择框，让用户挑备份保存位置（pywebview FileDialog）"""
        try:
            import webview
            w = webview.windows[0] if webview.windows else None
            if w is None:
                return {"ok": False, "error": "窗口不可用", "dir": None}
            try:
                dialog_type = webview.FileDialog.FOLDER
            except AttributeError:
                dialog_type = webview.FOLDER_DIALOG
            result = w.create_file_dialog(dialog_type)
            if not result:
                return {"ok": True, "dir": None}  # 用户取消
            d = result[0] if isinstance(result, (list, tuple)) else result
            return {"ok": True, "dir": d}
        except Exception as e:
            return {"ok": False, "error": str(e), "dir": None}

    def save_backup(self, directory, filename, content):
        """把备份 JSON 写入用户指定目录（2026-08-06 hermes：备份路径用户自选）"""
        try:
            if not directory or not filename:
                return {"ok": False, "error": "目录或文件名缺失"}
            path = os.path.join(directory, filename)
            with open(path, "w", encoding="utf-8") as f:
                f.write(content)
            return {"ok": True, "path": path}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ===== 团务档案/知识库双向打通 =====
    def list_archive_categories(self):
        """列出团务档案分类目录"""
        try:
            base = r"D:\小马儿-团务档案"
            if not os.path.isdir(base):
                return {"ok": False, "error": "团务档案目录不存在"}
            cats = []
            for name in sorted(os.listdir(base)):
                full = os.path.join(base, name)
                if os.path.isdir(full) and not name.startswith("_") and not name.startswith("."):
                    files = [f for f in os.listdir(full) if not f.startswith("~$")]
                    cats.append({"name": name, "count": len(files)})
            return {"ok": True, "categories": cats, "base": base}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def list_archive_templates(self, category=None):
        """列出团务档案参考模板文件"""
        try:
            base = r"D:\小马儿-团务档案\08-参考模板"
            if not os.path.isdir(base):
                return {"ok": False, "error": "模板目录不存在"}
            files = []
            for f in sorted(os.listdir(base)):
                if f.startswith("~$") or f.endswith(".docx"):
                    continue
                files.append({"name": f, "size": os.path.getsize(os.path.join(base, f))})
            return {"ok": True, "templates": files}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def read_template(self, filename):
        """读取模板内容（用于带入记录表单）"""
        try:
            base = r"D:\小马儿-团务档案\08-参考模板"
            safe = os.path.basename(filename)
            fpath = os.path.join(base, safe)
            if not os.path.exists(fpath):
                return {"ok": False, "error": "模板不存在"}
            with open(fpath, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()
            return {"ok": True, "content": content[:6000], "name": safe}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def archive_meeting(self, json_str, category="01-三会一课"):
        """把会议/活动记录归档到团务档案对应分类"""
        try:
            import time as _t
            data = json.loads(json_str)
            base = r"D:\小马儿-团务档案"
            cat_dir = os.path.join(base, category)
            if not os.path.isdir(cat_dir):
                os.makedirs(cat_dir, exist_ok=True)
            title = data.get("title") or data.get("name") or "工作记录"
            date = data.get("date") or _t.strftime("%Y-%m-%d")
            fname = f"{date}-{title[:40]}.md"
            fpath = os.path.join(cat_dir, fname)
            content = f"---\ntitle: {title}\ncreated: {date}\ntype: document\ncategories: [团务工作]\n---\n\n# {title}\n\n- 日期：{date}\n- 来源：团支书工作台自动归档\n\n## 📄 内容\n\n{data.get('content', '')}\n"
            with open(fpath, "w", encoding="utf-8") as f:
                f.write(content)
            return {"ok": True, "path": fpath, "filename": fname}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def generate_report(self, db_json):
        """基于工作台数据一键生成团支书述职报告"""
        try:
            db = json.loads(db_json) if isinstance(db_json, str) else db_json
            members = db.get("members", [])
            meetings = db.get("meetings", [])
            volunteers = db.get("volunteers", [])
            activities = db.get("activities", [])
            awards = db.get("awards", [])
            todos = db.get("todos", [])

            total = len(members)
            males = sum(1 for m in members if m.get("gender") == "男")
            females = sum(1 for m in members if m.get("gender") == "女")
            by_role = {}
            for m in members:
                r = m.get("role") or "普通团员"
                by_role[r] = by_role.get(r, 0) + 1
            role_desc = "、".join(f"{k}：{v}人" for k, v in by_role.items())

            def cnt(t):
                return len([m for m in meetings if m.get("type") == t])
            dahui = cnt("支部大会")
            zhivei = cnt("支委会")
            xiaozu = cnt("团小组会")
            tuanke = cnt("团课")

            vol_h = sum(float(v.get("hours", 0) or 0) for v in volunteers)
            vol_n = len(volunteers)
            act_n = len(activities)
            award_n = len(awards)
            todo_done = sum(1 for t in todos if t.get("done"))

            year = datetime.now().year
            report = f"""# {year}年度基层班级团支部工作述职报告

**2025级通信工程2班团支部书记 团支书**

## 引言

根据《中国共产主义青年团章程》、《中国共产主义青年团支部工作条例（试行）》以及学院团委相关通知要求，本人作为2025级通信工程2班团支部书记，在学院团委的直接领导和辅导员老师的指导下，切实履行支部建设第一责任人职责，团结带领全体团员青年，聚焦提升团支部引领力、组织力、服务力，扎实推进各项工作。现将本年度履职情况汇报如下。

## 一、组织情况

本支部现有成员 {total} 名，其中男生 {males} 名，女生 {females} 名。支部架构完整，职责明确：{role_desc}。

## 二、主要工作成效

（一）**严格落实"三会两制一课"**。本年度主持召开支部大会 {dahui} 次、支委会 {zhivei} 次、团小组会 {xiaozu} 次，组织团课 {tuanke} 次，内容涵盖工作部署、议题研讨、理论学习等，确保组织生活严肃有序。

（二）**扎实推进思想引领与实践育人**。组织劳动实践 {vol_n} 次，累计 {vol_h} 小时；开展班级活动 {act_n} 次，将思想教育融入具体实践。

（三）**规范团员管理与评优工作**。做好团员台账动态维护、团费收缴、推优入党材料准备；本年度参与评优评先 {award_n} 项，完成待办事项 {todo_done} 项。

## 三、存在不足

1. 理论学习深度有待加强，部分活动形式较为单一；
2. 团员参与积极性需进一步调动；
3. 工作留痕与档案归档仍需规范。

## 四、未来展望

新的一年，我将继续以服务同学为宗旨，创新活动形式，加强支部建设，提升团支部的凝聚力与战斗力，为学院共青团工作贡献青春力量。

2025级通信工程2班团支部
{year}年12月
"""
            return {"ok": True, "report": report}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def batch_archive(self, db_json):
        """一键归档：把工作台全部记录自动分类写入团务档案（自动去重）"""
        try:
            import time as _t
            db = json.loads(db_json) if isinstance(db_json, str) else db_json
            base = r"D:\小马儿-团务档案"
            if not os.path.isdir(base):
                return {"ok": False, "error": "团务档案目录不存在"}

            stats = {}   # category -> {added, skipped}
            results = []

            def write_record(cat, title, date, content):
                cat_dir = os.path.join(base, cat)
                if not os.path.isdir(cat_dir):
                    os.makedirs(cat_dir, exist_ok=True)
                fname = f"{date}-{title[:40]}.md"
                fpath = os.path.join(cat_dir, fname)
                if os.path.exists(fpath):
                    stats.setdefault(cat, {"added": 0, "skipped": 0})["skipped"] += 1
                    return False
                md = f"---\ntitle: {title}\ncreated: {date}\ntype: document\ncategories: [团务工作]\nsource: 团支书工作台一键归档\n---\n\n# {title}\n\n- 日期：{date}\n- 来源：团支书工作台一键归档\n\n## 📄 内容\n\n{content}\n"
                with open(fpath, "w", encoding="utf-8") as f:
                    f.write(md)
                stats.setdefault(cat, {"added": 0, "skipped": 0})["added"] += 1
                results.append(fpath)
                return True

            # 1. 会议记录 → 01-三会一课（按类型标注）
            for m in db.get("meetings", []):
                title = f"{m.get('type','会议')}-{m.get('name','')[:28]}"
                content = f"会议类型：{m.get('type')}\n时间：{m.get('date')}\n主题：{m.get('name')}\n内容：{m.get('content') or m.get('note') or ''}"
                write_record("01-三会一课", title, m.get("date") or _t.strftime("%Y-%m-%d"), content)

            # 2. 活动 → 02-团日活动
            for a in db.get("activities", []):
                title = f"活动-{a.get('title','')[:28]}"
                date = (a.get("startTime") or "").replace("T", " ")[:10] or _t.strftime("%Y-%m-%d")
                content = f"活动时间：{a.get('startTime') or ''} → {a.get('endTime') or ''}\n活动名称：{a.get('title')}\n活动要求：{a.get('requirement') or ''}\n参加人数：{len(a.get('attendees') or [])} 人\n参加人员：{'、'.join(a.get('attendees') or [])}"
                write_record("02-团日活动", title, date, content)

            # 3. 班级活动 → 03-主题班会
            for c in db.get("classActs", []):
                title = f"班会-{c.get('name','')[:28]}"
                content = f"时间：{c.get('date')}\n活动：{c.get('name')}\n地点：{c.get('place')}\n参与人数：{c.get('attendees')}\n内容：{c.get('desc') or ''}"
                write_record("03-主题班会", title, c.get("date") or _t.strftime("%Y-%m-%d"), content)

            # 4. 通知 → 04-班级数据
            for n in db.get("notices", []):
                title = f"通知-{n.get('title','')[:28]}"
                content = f"发布时间：{n.get('date')}\n渠道：{n.get('channel')}\n状态：{n.get('status')}\n内容：{n.get('content') or ''}"
                write_record("04-班级数据", title, n.get("date") or _t.strftime("%Y-%m-%d"), content)

            # 5. 评优 → 06-党建述职
            for aw in db.get("awards", []):
                title = f"评优-{aw.get('name','')[:28]}"
                content = f"时间：{aw.get('date')}\n奖项：{aw.get('name')}\n详情：{aw.get('note') or aw.get('desc') or ''}"
                write_record("06-党建述职", title, aw.get("date") or _t.strftime("%Y-%m-%d"), content)

            # 汇总
            summary = {k: v for k, v in stats.items()}
            total_added = sum(v["added"] for v in stats.values())
            total_skipped = sum(v["skipped"] for v in stats.values())
            return {"ok": True, "summary": summary, "added": total_added, "skipped": total_skipped, "files": results}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def export_md(self, filename, content):
        """导出 markdown 文件到桌面"""
        try:
            import time as _t
            desktop = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
            if not os.path.isdir(desktop):
                desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            safe = os.path.basename(filename).replace(" ", "_")
            if not safe.endswith(".md"):
                safe += ".md"
            fpath = os.path.join(desktop, safe)
            with open(fpath, "w", encoding="utf-8") as f:
                f.write(content)
            return {"ok": True, "path": fpath}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def open_browser(self, url):
        """在系统默认浏览器打开外部链接（智慧团建等）"""
        try:
            webbrowser.open(url)
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def import_excel(self, file_path):
        """从 Excel 文件导入团员信息
        支持的列：姓名/学号/性别/民族/手机号码/团内职务/团籍是否在本组织/入团年月
        """
        if not HAS_OPENPYXL:
            return {"ok": False, "error": "缺少 openpyxl 库，请运行: pip install openpyxl"}
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return {"ok": False, "error": "Excel 为空"}
            # 表头映射
            header = [str(c).strip() if c else "" for c in rows[0]]
            col_map = {}
            dorm_building = dorm_room = None
            for i, h in enumerate(header):
                key = None
                if "姓名" in h: key = "name"
                elif "学号" in h or "序号" in h: key = "studentId" if "学号" in h else "seq"
                elif "性别" in h: key = "gender"
                elif "民族" in h: key = "ethnicity"
                elif "手机" in h or "电话" in h: key = "phone"
                elif "职务" in h: key = "role"
                elif "团籍" in h or "组织" in h: key = "inOrg"
                elif "入团" in h or "团年月" in h: key = "joinDate"
                elif "楼" in h or "栋" in h or "宿舍楼" in h: dorm_building = i; key = None
                elif "宿舍" in h or "房间" in h or "房号" in h or "寝室" in h: dorm_room = i; key = None
                if key:
                    col_map[i] = key
            # 表头缺失时按位置推断（宿舍名单常见格式：班级|学号|姓名|性别|楼栋|宿舍号）
            if "gender" not in col_map.values():
                # 常见布局：姓名列后紧跟性别
                name_idx = [i for i, k in col_map.items() if k == "name"]
                if name_idx:
                    ni = name_idx[0]
                    # 性别 = 姓名后第1列（若为空表头）
                    if ni + 1 < len(header) and not header[ni + 1]:
                        col_map[ni + 1] = "gender"
                        # 楼栋 = 性别后第1列，宿舍 = 再后1列
                        if ni + 2 < len(header) and not header[ni + 2]:
                            dorm_building = ni + 2
                        if ni + 3 < len(header) and not header[ni + 3]:
                            dorm_room = ni + 3
            if "name" not in col_map.values():
                return {"ok": False, "error": "未找到'姓名'列"}
            # 解析数据行
            members = []
            for row in rows[1:]:
                if not row or not any(c is not None and str(c).strip() for c in row):
                    continue
                m = {}
                for i, key in col_map.items():
                    val = row[i] if i < len(row) else None
                    if val is None:
                        m[key] = ""
                        continue
                    if key == "joinDate":
                        m[key] = _norm_join_date(val)
                    elif key == "inOrg":
                        s = str(val).strip()
                        m[key] = "是" if s in ("是", "Y", "y", "yes", "YES", "1", "true", "True") else ("否" if s else "")
                    elif key == "seq":
                        pass  # 序号忽略
                    else:
                        m[key] = str(val).strip()
                if m.get("name"):
                    # 宿舍合并：楼栋+房间 → "B00#000"
                    if dorm_building is not None or dorm_room is not None:
                        b = str(row[dorm_building]).strip() if dorm_building is not None and dorm_building < len(row) and row[dorm_building] else ""
                        r = str(row[dorm_room]).strip() if dorm_room is not None and dorm_room < len(row) and row[dorm_room] else ""
                        b = b.replace("栋", "").replace("楼", "").replace("B", "B").strip()
                        if b and r:
                            m["dorm"] = f"{b}#{r}"
                        elif r:
                            m["dorm"] = r
                    members.append(m)
            return {"ok": True, "members": members, "count": len(members)}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def parse_excel_base64(self, b64_str, filename="团员列表.xlsx"):
        """前端传 base64 的 Excel 文件，解析团员信息"""
        import base64 as b64mod
        import tempfile
        try:
            raw = b64mod.b64decode(b64_str)
            tmp_path = os.path.join(tempfile.gettempdir(), "tuan_import_" + filename)
            with open(tmp_path, "wb") as f:
                f.write(raw)
            return self.import_excel(tmp_path)
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def save_photo(self, b64_str, filename="photo.jpg"):
        """保存劳动实践照片到应用目录 photos/ 文件夹"""
        import base64 as b64mod
        import time
        try:
            photos_dir = os.path.join(APP_DIR, "photos")
            os.makedirs(photos_dir, exist_ok=True)
            raw = b64mod.b64decode(b64_str)
            # 安全文件名
            safe = os.path.basename(filename).replace(' ', '_')
            name, ext = os.path.splitext(safe)
            ext = (ext or '.jpg').lower()
            if ext not in ('.jpg', '.jpeg', '.png', '.gif', '.webp'):
                ext = '.jpg'
            fname = f"{int(time.time())}_{name[:20]}{ext}"
            fpath = os.path.join(photos_dir, fname)
            with open(fpath, "wb") as f:
                f.write(raw)
            return {"ok": True, "filename": fname, "path": fpath}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def list_photos(self):
        """列出 photos/ 文件夹里的所有照片"""
        try:
            photos_dir = os.path.join(APP_DIR, "photos")
            if not os.path.isdir(photos_dir):
                return {"ok": True, "photos": []}
            files = sorted(os.listdir(photos_dir))
            photos = []
            for f in files:
                if f.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp')):
                    fpath = os.path.join(photos_dir, f)
                    photos.append({
                        "filename": f,
                        "size": os.path.getsize(fpath),
                        "mtime": datetime.fromtimestamp(os.path.getmtime(fpath)).strftime("%Y-%m-%d %H:%M")
                    })
            return {"ok": True, "photos": photos}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def delete_photo(self, filename):
        """删除照片文件"""
        try:
            photos_dir = os.path.join(APP_DIR, "photos")
            safe = os.path.basename(filename)
            fpath = os.path.join(photos_dir, safe)
            if os.path.exists(fpath):
                os.remove(fpath)
                return {"ok": True}
            return {"ok": False, "error": "文件不存在"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def save_activity_file(self, b64_str, filename="activity.txt"):
        """保存活动文件到 activities/ 文件夹"""
        import base64 as b64mod
        import time
        try:
            act_dir = os.path.join(APP_DIR, "activities")
            os.makedirs(act_dir, exist_ok=True)
            raw = b64mod.b64decode(b64_str)
            safe = os.path.basename(filename).replace(' ', '_')
            name, ext = os.path.splitext(safe)
            ext = (ext or '.txt').lower()
            if len(ext) > 8 or not ext[1:].isalnum():
                ext = '.txt'
            fname = f"{int(time.time())}_{name[:24]}{ext}"
            fpath = os.path.join(act_dir, fname)
            with open(fpath, "wb") as f:
                f.write(raw)
            return {"ok": True, "filename": fname, "path": fpath}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def list_activity_files(self):
        """列出 activities/ 文件夹里的所有文件"""
        try:
            act_dir = os.path.join(APP_DIR, "activities")
            if not os.path.isdir(act_dir):
                return {"ok": True, "files": []}
            files = []
            for f in sorted(os.listdir(act_dir)):
                fpath = os.path.join(act_dir, f)
                if os.path.isfile(fpath):
                    files.append({
                        "filename": f,
                        "size": os.path.getsize(fpath),
                        "mtime": datetime.fromtimestamp(os.path.getmtime(fpath)).strftime("%Y-%m-%d %H:%M")
                    })
            return {"ok": True, "files": files}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def delete_activity_file(self, filename):
        """删除活动文件"""
        try:
            act_dir = os.path.join(APP_DIR, "activities")
            safe = os.path.basename(filename)
            fpath = os.path.join(act_dir, safe)
            if os.path.exists(fpath):
                os.remove(fpath)
                return {"ok": True}
            return {"ok": False, "error": "文件不存在"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def export_report(self, json_str):
        """导出完整报表"""
        try:
            path = os.path.join(APP_DIR, f"工作台导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
            with open(path, "w", encoding="utf-8") as f:
                f.write(json_str)
            return {"ok": True, "path": path}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def ask_ai(self, user_msg, db_json):
        """调用 DeepSeek API 作为工作台内嵌 AI 助手（团务助手），能读取工作台数据"""
        try:
            # 读取 API key（DEEPSEEK_API_KEY）
            api_key = os.environ.get("DEEPSEEK_API_KEY", "")
            if not api_key:
                env_path = os.path.join(os.path.expanduser("~"), "AppData", "Local", "hermes", ".env")
                if os.path.exists(env_path):
                    with open(env_path, "r", encoding="utf-8") as f:
                        for line in f:
                            if line.startswith("DEEPSEEK_API_KEY="):
                                api_key = line.strip().split("=", 1)[1]
                                break
            if not api_key:
                return {"ok": False, "error": "未找到 DEEPSEEK_API_KEY"}

            # 解析工作台数据，生成精简摘要上下文
            summary = self._summarize_db(db_json)

            system_prompt = (
                "你是嵌入在'班团一体化工作台'里的 AI 助手，名叫团务。"
                "你的用户是高校团支书团支书（某高校25通信工程X班，搭档班长班长）。"
                "你可以看到他工作台里的实时数据（见下方【工作台数据摘要】），帮他分析团务、班级事务、活动、考勤、劳动实践等。"
                "回答简洁务实、直接给结论，可以用列表/要点。不要编造数据，只基于给定数据回答。"
                "如果数据里没有的，明确说'工作台里还没有这个数据'。\n\n"
                f"【工作台数据摘要】\n{summary}"
            )

            body = json.dumps({
                "model": "deepseek-chat",
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_msg}
                ],
                "temperature": 0.7,
                "max_tokens": 1500
            }).encode("utf-8")

            req = urllib.request.Request(
                "https://api.deepseek.com/v1/chat/completions",
                data=body,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": "Bearer " + api_key
                }
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                reply = data["choices"][0]["message"]["content"]
                return {"ok": True, "reply": reply}
        except urllib.error.HTTPError as e:
            # 接口错误：读响应体给出具体原因
            try:
                err_body = e.read().decode("utf-8")
            except Exception:
                err_body = ""
            return {"ok": False, "error": f"AI接口HTTP {e.code}: {err_body[:200]}"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def open_ai_window(self):
        """新开独立窗口显示 AI 助手"""
        try:
            import webview
            # 复用同一个 api（共享数据）
            api = self
            ai_path = os.path.join(APP_DIR, "ai.html")
            w2 = webview.create_window(
                "🤖 团务助手 · 班团工作台",
                ai_path,
                js_api=api,
                width=520,
                height=700,
                min_size=(420, 520),
                background_color=window_bg_color(),
                text_select=True,
            )
            w2.events.loaded += lambda: None
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def open_smart_window(self, url=None):
        """新开独立窗口显示智慧团建官网（外部网站，不注入 js_api 防安全风险）"""
        try:
            import webview
            target = url or "https://zhtj.youth.cn/zhtj/signin"
            w2 = webview.create_window(
                "智慧团建 · 团中央系统",
                target,
                width=1180,
                height=820,
                min_size=(900, 600),
                text_select=True,
            )
            w2.events.loaded += lambda: None
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def nav_zhtj(self):
        """主窗口直接导航到智慧团建官网（第一方上下文，登录态可保持）"""
        try:
            import webview
            for w in webview.windows:
                if w is not None and hasattr(w, "load_url"):
                    w.load_url("https://zhtj.youth.cn/zhtj/signin")
                    return {"ok": True}
            return {"ok": False, "error": "主窗口未找到"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def nav_back(self):
        """主窗口导航回工作台首页（用完智慧团建后返回）"""
        try:
            import webview
            index_path = os.path.join(APP_DIR, "index.html")
            for w in webview.windows:
                if w is not None and hasattr(w, "load_url"):
                    w.load_url(index_path)
                    return {"ok": True}
            return {"ok": False, "error": "主窗口未找到"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def parse_docx(self, b64, filename=None):
        """解析 Word 文档（.docx），提取正文文本（zipfile + XML，标准库零依赖）"""
        try:
            import base64, io, zipfile
            import xml.etree.ElementTree as ET
            raw = base64.b64decode(b64)
            zf = zipfile.ZipFile(io.BytesIO(raw))
            if "word/document.xml" not in zf.namelist():
                return {"ok": False, "error": "不是有效的 .docx 文件（老版 .doc 格式不支持）"}
            xml_data = zf.read("word/document.xml")
            W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
            root = ET.fromstring(xml_data)
            paras = []
            for p in root.iter(W + "p"):
                texts = [t.text or "" for t in p.iter(W + "t")]
                line = "".join(texts).strip()
                if line:
                    paras.append(line)
            text = "\n".join(paras)
            title = (filename or "会议记录")
            import os as _os
            for ext in (".docx", ".doc"):
                if title.lower().endswith(ext):
                    title = title[: -len(ext)]
            return {"ok": True, "title": title, "text": text[:5000]}
        except Exception as e:
            return {"ok": False, "error": "解析失败：" + str(e)}

    def _summarize_db(self, db_json):
        """把工作台数据压缩成 AI 可读的摘要"""
        try:
            db = json.loads(db_json) if isinstance(db_json, str) else db_json
        except Exception:
            return "（数据解析失败）"
        lines = []
        members = db.get("members", [])
        lines.append(f"团员/学生：{len(members)} 人")
        if members:
            roles = {}
            for m in members:
                if m.get("role"):
                    roles[m["role"]] = roles.get(m["role"], 0) + 1
            if roles:
                lines.append("  职务分布：" + "、".join(f"{k}{v}人" for k, v in roles.items()))
        vols = db.get("volunteers", [])
        if vols:
            total_h = sum(float(v.get("hours", 0) or 0) for v in vols)
            lines.append(f"劳动实践：{len(vols)} 条记录，共 {total_h} 小时")
        atts = db.get("attendance", [])
        if atts:
            from collections import Counter
            c = Counter(a.get("type", "?") for a in atts)
            lines.append("考勤：" + "、".join(f"{k}{v}" for k, v in c.items()))
        acts = db.get("activities", [])
        if acts:
            lines.append(f"活动：{len(acts)} 个（进行中/未开始 {sum(1 for a in acts if not a.get('endTime') or a['endTime'] >= datetime.now().isoformat()[:16])} 个）")
        meets = db.get("meetings", [])
        if meets:
            from collections import Counter
            c = Counter(m.get("type", "?") for m in meets)
            lines.append("组织生活：" + "、".join(f"{k}{v}次" for k, v in c.items()))
        todos = db.get("todos", [])
        if todos:
            lines.append(f"待办：共 {len(todos)} 项，未完成 {sum(1 for t in todos if not t.get('done'))} 项")
        notices = db.get("notices", [])
        if notices:
            lines.append(f"通知：{len(notices)} 条")
        return "\n".join(lines) if lines else "（工作台暂无数据）"


def main():
    import webview

    # 窗口大小：贴边时占页面总宽度 3/5（不占 1/2）
    try:
        import ctypes
        user32 = ctypes.windll.user32
        sw = user32.GetSystemMetrics(0)
        sh = user32.GetSystemMetrics(1)
        w = int(sw * 3 / 5)  # 3/5 宽
        h = sh - 80  # 高度留出任务栏
    except Exception:
        w, h = 1150, 800

    api = Api()
    index_path = os.path.join(APP_DIR, "index.html")

    # 启动时自动备份数据（每天一个快照，当天已有则跳过；防数据丢失）
    try:
        data_path = os.path.join(APP_DIR, "tuanzhishu_data.json")
        if os.path.exists(data_path):
            bdir = os.path.join(APP_DIR, ".backup")
            os.makedirs(bdir, exist_ok=True)
            today = datetime.now().strftime("%Y-%m-%d")
            target = os.path.join(bdir, f"tuanzhishu_{today}.json")
            if not os.path.exists(target):
                import shutil
                shutil.copy2(data_path, target)
    except Exception:
        pass

    window = webview.create_window(
        "🇨🇳 团支书工作台 · 某高校 25通信工程X班",
        index_path,
        js_api=api,
        width=w,
        height=h,
        min_size=(720, 540),
        background_color=window_bg_color(),
        text_select=True,
    )

    # 窗口加载后直接把数据注入前端（绕开 pywebview 桥接中文编码问题）
    def inject_data():
        try:
            data_path = os.path.join(APP_DIR, "tuanzhishu_data.json")
            if os.path.exists(data_path):
                with open(data_path, "r", encoding="utf-8") as f:
                    raw = f.read()
                import json as _json
                window.evaluate_js(f"window.__injected_data = {_json.dumps(raw)};")
        except Exception:
            pass
    window.events.loaded += inject_data

    webview.start(debug=False, storage_path=os.path.join(APP_DIR, ".webview_profile"))

if __name__ == "__main__":
    main()
